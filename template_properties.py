# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import *
from openpyxl.worksheet.worksheet import *
from openpyxl.utils import get_column_letter
from BMS_template.utilities import DesignUtils
from BMS_template.write_attributes import AttributesToTemplate

class Properties:

    def template_lang(lang):

        if lang == 'UA':
            template_lang = 0
        elif lang == 'RU':
            template_lang = 1
        elif lang == 'EU':
            template_lang = 2
        else:
            template_lang = None
            print('Wrong language')
        return template_lang

    def template_type(lang):
        template_lang = Properties.template_lang(lang)
        sheet_names_en = ['BvD Results','Search Strategy','BvD Strategy Steps','Manual steps','Search Matrix', 'Web review','Market range','Reconciliation matrix']
        sheet_names = AttributesToTemplate.get_attributes(template_lang, sheet_names_en)
        return sheet_names

    def lang_from_template_type(bms_template):
        sheet_names = bms_template.get_sheet_names()
        lang = 'Undetermined'
        if sheet_names == list(Properties.template_type('UA').values()):
            lang = 'UA'
        elif sheet_names == list(Properties.template_type('RU').values()):
            lang = 'RU'
        elif sheet_names == list(Properties.template_type('EU').values()):
            lang = 'EU'
        return lang

    def define_col_width(sheet,col_range,etalon_row):

        empty_col_width = 6
        max_col_width = 16
        for col in range(col_range[0],col_range[1]):
            cell_content = str(sheet.cell(column=col,row = etalon_row).value)
            if cell_content is None:
                sheet.column_dimensions[get_column_letter(col)].width = empty_col_width
            word_lengths = []
            word_count = 0
            for word in cell_content.split():
                word_count = word_count+1
                word_lengths.append(len(word))
            if max(word_lengths)<= empty_col_width and word_count == 1:
                sheet.column_dimensions[get_column_letter(col)].width = empty_col_width
            elif max(word_lengths)*1.5 > max_col_width and word_count <= 8:
                sheet.column_dimensions[get_column_letter(col)].width = max_col_width
            elif word_count <= 8 or not (word_count == 1 and word_lengths[0]*1.5 <= 3):
                sheet.column_dimensions[get_column_letter(col)].width = max(word_lengths)*1.5
            elif word_count>=8:
                sheet.column_dimensions[get_column_letter(col)].width = max_col_width*2
            else:
                sheet.column_dimensions[get_column_letter(col)].width = empty_col_width
        return sheet

    def number_to_format(sheet,column,row, format_str):
        cell = sheet.cell(column=column, row=row)
        try:
            n = float(cell.value)
        except ValueError:
            n = cell.value
        _ = sheet.cell(column=column,row=row, value = n)
        sheet.cell(column=column, row=row).number_format = format_str
        return sheet


class Styles:
    def basic_style(cell):

        cell.font = Font(name='Verdana', size=8)
        cell.border = Border(left=Side(border_style='thin',color=colors.BLACK),
                      right=Side(border_style='thin',color=colors.BLACK),
                      top=Side(border_style='thin',color=colors.BLACK),
                      bottom=Side(border_style='thin',color=colors.BLACK))
        cell.alignment = Alignment(horizontal='left',vertical='bottom',wrap_text=False)
        return cell


    def header_style_dark_blue(cell):

        cell.font = Font(name='Verdana', size=8, bold=True, color=colors.WHITE)
        cell.border = Border(left=Side(border_style='thin', color=colors.WHITE),
                             right=Side(border_style='thin', color=colors.WHITE),
                             top=Side(border_style='thin', color=colors.WHITE),
                             bottom=Side(border_style='thin', color=colors.WHITE))
        cell.fill = PatternFill(patternType='solid',fgColor=DesignUtils.color_palette()['BLUE6'])
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        return cell

    def header_style_green(cell):

        cell.font = Font(name='Verdana', size=8, bold=True, color=colors.WHITE)
        cell.border = Border(left=Side(border_style='thin', color=colors.WHITE),
                             right=Side(border_style='thin', color=colors.WHITE),
                             top=Side(border_style='thin', color=colors.WHITE),
                             bottom=Side(border_style='thin', color=colors.WHITE))
        cell.fill = PatternFill(patternType='solid',fgColor=DesignUtils.color_palette()['DELOITTEGREEN'])
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        return cell

    def header_style_light_blue(cell):

        cell.font = Font(name='Verdana', size=8, bold=True, color=colors.WHITE)
        cell.border = Border(left=Side(border_style='thin', color=colors.WHITE),
                             right=Side(border_style='thin', color=colors.WHITE),
                             top=Side(border_style='thin', color=colors.WHITE),
                             bottom=Side(border_style='thin', color=colors.WHITE))
        cell.fill = PatternFill(patternType='solid',fgColor=DesignUtils.color_palette()['BLUE2'])
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        return cell

    def search_strategy_style(cell):

        cell.font = Font(name='Verdana', size=8, bold=True, color=DesignUtils.color_palette()['BLUE6'])
        cell.border = Border(left=Side(border_style='thin', color=DesignUtils.color_palette()['BLUE6']),
                             right=Side(border_style='thin', color=DesignUtils.color_palette()['BLUE6']),
                             top=Side(border_style='thin', color=DesignUtils.color_palette()['BLUE6']),
                             bottom=Side(border_style='thin', color=DesignUtils.color_palette()['BLUE6']))
        cell.fill = PatternFill(patternType='solid', fgColor=DesignUtils.color_palette()['LIGHTBERLINBLUE'])
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        return cell

    def thick_border_white_cell_style(cell):

        cell.font = Font(name='Verdana', size=8, bold=True, color=DesignUtils.color_palette()['BLUE6'])
        cell.border = Border(left=Side(border_style='thick', color=DesignUtils.color_palette()['BLUE6']),
                             right=Side(border_style='thick', color=DesignUtils.color_palette()['BLUE6']),
                             top=Side(border_style='thick', color=DesignUtils.color_palette()['BLUE6']),
                             bottom=Side(border_style='thick', color=DesignUtils.color_palette()['BLUE6']))
        cell.alignment = Alignment(horizontal='center',vertical='center')

        return cell

    def thick_border_fill_cell_style(cell):

        cell.font = Font(name='Verdana', size=8, bold=True, color=DesignUtils.color_palette()['BLUE6'])
        cell.border = Border(left=Side(border_style='thick', color=DesignUtils.color_palette()['BLUE6']),
                             right=Side(border_style='thick', color=DesignUtils.color_palette()['BLUE6']),
                             top=Side(border_style='thick', color=DesignUtils.color_palette()['BLUE6']),
                             bottom=Side(border_style='thick', color=DesignUtils.color_palette()['BLUE6']))
        cell.fill = PatternFill(patternType='solid',fgColor=DesignUtils.color_palette()['LIGHTBERLINBLUE'])
        cell.alignment = Alignment(horizontal='center',vertical='center')

        return cell

    def thin_border_fill_cell_style(cell):

        cell.font = Font(name='Verdana', size=8, bold=True, color=DesignUtils.color_palette()['BLUE6'])
        cell.border = Border(left=Side(border_style='thin', color=DesignUtils.color_palette()['BLUE6']),
                             right=Side(border_style='thin', color=DesignUtils.color_palette()['BLUE6']),
                             top=Side(border_style='thin', color=DesignUtils.color_palette()['BLUE6']),
                             bottom=Side(border_style='thin', color=DesignUtils.color_palette()['BLUE6']))
        cell.fill = PatternFill(patternType='solid',fgColor=DesignUtils.color_palette()['LIGHTBERLINBLUE'])
        cell.alignment = Alignment(horizontal='left',vertical='center',wrap_text=True)

        return cell