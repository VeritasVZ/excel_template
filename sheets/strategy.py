from openpyxl.worksheet.worksheet import *
from BMS_template.template_properties import Properties, Styles


class StrategySheet:

    def write_strategy_steps(lang,strategy_sheet):
        total = ['Загалом','Всего','Total in BvD']
        accepted = ['Прийнято','Принято','Accepted']
        rejected = ['Відхилено','Отклонено','Rejected']

        if lang == 'UA':
            take = 0
        elif lang == 'RU':
            take = 1
        else: take = 2

        strategy_sheet['D1'] = total[take]
        strategy_sheet['E1'] = accepted[take]
        strategy_sheet['F1'] = rejected[take]

        for col in range(4,8):
            for row in range(1,16):
                try:
                    if strategy_sheet['D' + str(row)].value is not None:
                        temp = str(strategy_sheet['D' + str(row)].value)
                        temp = temp.split(',')
                        new_value = ''.join(temp)
                        strategy_sheet['D' + str(row)] = new_value
                        Properties.number_to_format(strategy_sheet, column=col, row=row, format_str='0')
                except TypeError or AttributeError:
                    continue
        return strategy_sheet


    def write_search_strategy(bms_template,search_strategy,lang): # bms_template --> excel file - result of write_bvd_results function, search strategy -->excel sheet
        sheet_names = bms_template.get_sheet_names()
        sheet = bms_template.get_sheet_by_name(sheet_names[1])

        for row in range(1,8):
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            sheet.row_dimensions[row].height = 10
        for row in range(8,16):
            sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        sheet.merge_cells(start_row = 16, start_column = 1, end_row = 16, end_column = 5)
        sheet.merge_cells(start_row=17, start_column=1, end_row=17, end_column=3)
        for row in range(18,22):
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

        for col in range(1,len(tuple(search_strategy.columns))+2):
            for row in range(1,len(tuple(search_strategy.rows))+1):
                cell_value = search_strategy.cell(column=col,row=row).value
                cell_in_template = sheet.cell(column=col,row = row, value = cell_value)
                Styles.search_strategy_style(cell_in_template)

        for row in range(8, 16):
            lenght = len(sheet.cell(column=2, row=row).value)
            sheet.row_dimensions[row].height = 10*(lenght/100+1)

        sheet.column_dimensions[get_column_letter(1)].width = 3
        sheet.column_dimensions[get_column_letter(2)].width = 70
        sheet.column_dimensions[get_column_letter(3)].width = 18
        for col in range (4,8):
            sheet.column_dimensions[get_column_letter(col)].width = 12
        sheet = StrategySheet.write_strategy_steps(lang=lang,strategy_sheet=sheet)

        print('strategy added')
        return bms_template
