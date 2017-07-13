from openpyxl.styles import *
from openpyxl.drawing.image import Image, bounding_box
from openpyxl.worksheet.worksheet import *
from openpyxl.utils import get_column_letter
from BMS_template.utilities import DesignUtils
from BMS_template.template_properties import Properties, Styles


class ManualStep:

    def define_step_name_type(lang):

        step_name_types = {
            'criterion 1': ('Відбір за виручкою\nМінімальна середня виручка (тис) = \nМаксимальна середня виручка (тис) = ',
                            'Отбор по выручке\nМинимальная средняя выручка  (тыс) = \nМаксимальная средняя выручка (тыс) = ',
                            'Revenue threshold\nMinimal average revenue (th) = \nMaximal average revenue (th) = '),
            'criterion 2': ('Відбір за прибутком\nМінімальний прибуток (тис) = ', 'Отбор по прибыли\nМинимальная прибыль (тыс) = ', 'Net Profit threshold\nMinimal Net profit (th) = '),
            'criterion 3': ('Вибірка для перевірки в інтернеті', 'Выборка для проверки в интернете', 'Web search'),
            'criterion 4': ('Відбір за доступністю інформації', 'Отбор по доступности информации', 'Information availability'),
            'criterion 5': ('Відбір за предметом операцій', 'Отбор по предмету операций', 'Product comparability'),
            'criterion 6': ('Відбір за зіставністю функцій', 'Отбор по функциональному профилю', 'Functions comparability'),
            'criterion 7': ('Відбір за наявністю додаткової незіставної діяльності', 'Отбор по наличию дополнительной несопоставимой деятельности','Additional uncomparable functions'),
            'criterion 8': ('Відбір за критерієм незалежності', 'Отбор по критерию независимости', 'Independence check'),
            'criterion 9': ('Вибірка зіставних компаній', 'Выборка сопоставимых компаний', 'Final comparable set'),
            'criterion 10': ('Результати дослідження діяльності обраних компаній за інформацією, отриманою з мережі Інтернет, та додаткової перевірки незалежності компаній',
                             'Результаты исследования деятельности отобранных компаний по информации полученной из открытого источника Интернет, и дополнительной проверки независимости сопоставляемых компаний',
                             'Results of comparable companies review in Internet and independence doublecheck')
        }
        if lang == 'UA':
            step_name_type = 0
        elif lang == 'RU':
            step_name_type = 1
        elif lang == 'EU':
            step_name_type = 2
        else:
            print('Wrong language')

        step_name_type_list = []
        for criterion in list(step_name_types.keys()):
            step_name_type_list.append(step_name_types.get(criterion)[step_name_type])

        return step_name_type_list

    def steps_table(bms_template):

        sheet_names = bms_template.get_sheet_names()
        step_sheet = bms_template.get_sheet_by_name(sheet_names[3])
        img = Image('Deloitte_logo.png',size=[250,50])
        img.anchor(step_sheet.cell('B1'),anchortype='oneCell')
        step_sheet.add_image(img)

        step_sheet.merge_cells(start_row=1, start_column=1, end_row=3, end_column=5)
        for col in range(1,6):
            for row in range(1,4):
                Styles.thick_border_white_cell_style(step_sheet.cell(column=col,row=row))
        step_sheet.column_dimensions[get_column_letter(1)].width = 3
        step_sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=3)
        for col in range (1,6):
            Styles.thick_border_fill_cell_style(step_sheet.cell(column=col,row=4))
        step_sheet.column_dimensions[get_column_letter(2)].width = 35
        step_sheet.column_dimensions[get_column_letter(3)].width = 10
        step_sheet.column_dimensions[get_column_letter(4)].width = 15
        step_sheet.column_dimensions[get_column_letter(5)].width = 15

        for row in range(1,5):
            step_sheet.row_dimensions[row].height = 13

        lang = Properties.lang_from_template_type(bms_template)
        for col in range(1, 6):
            Styles.thin_border_fill_cell_style(step_sheet.cell(column=col, row=5))
            step_sheet.row_dimensions[row].height = 50
            step_sheet.cell(column=2, row=5, value=ManualStep.define_step_name_type(lang)[0])
            Styles.thin_border_fill_cell_style(step_sheet.cell(column=col, row=6))
            step_sheet.row_dimensions[row].height = 30
            step_sheet.cell(column=2, row=6, value=ManualStep.define_step_name_type(lang)[1])

        step_sheet['A5'] = 1
        step_sheet['A6'] = 2
        for row in range(7,14):
            step_sheet['A'+str(row)]=row-4
            step_sheet.row_dimensions[row].height = 20
            step_sheet.cell(column=2, row=row, value=ManualStep.define_step_name_type(lang)[row-5])
            step_sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            for col in range(1,6):
                Styles.thin_border_fill_cell_style(step_sheet.cell(column=col,row=row))

        step_sheet.row_dimensions[14].height = 30
        for col in range (1,6):
            Styles.header_style_light_blue(step_sheet.cell(column=col,row=14))
        step_sheet['B14'] = ManualStep.define_step_name_type(lang)[9]
        step_sheet.merge_cells(start_row=14, start_column=2, end_row=14, end_column=3)

        print('manual steps added')
        return bms_template