import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import *
from template_properties import Properties, Styles

class BvDResultsSheet:

    def write_bvd_results(bms_template,bvd_results_df):  # bms_template --> empty excel file; bms_results-->pandas dataframe
        sheet_names = bms_template.get_sheet_names()
        sheet = bms_template.get_sheet_by_name(sheet_names[0])
        for row in dataframe_to_rows(bvd_results_df, index=True, header=True):
            sheet.append(row)
        col_range = (1,len(sheet[1])+1)
        Properties.define_col_width(sheet,col_range,2)

        for col in range(1, len(bvd_results_df.columns) + 2):
            cell = sheet.cell(column=col, row=1)
            Styles.header_style_dark_blue(cell)
            sheet.row_dimensions[1].height = 50
            for row in range(2, len(bvd_results_df.index) + 1):
                cell = sheet.cell(column=col, row=row)
                Styles.basic_style(cell)

        for col in range(BvDResultsSheet.first_fin_data_row(bvd_results_df)+2,len(bvd_results_df.columns ) +2):
            sheet.column_dimensions[get_column_letter(col)].width = 20
            for row in range(2, len(bvd_results_df.index ) +1):
                Properties.number_to_format(sheet ,column=col ,row=row ,format_str='0.0')
        print('bvd_results added')
        return bms_template

    def first_fin_data_row(bvd_results_df):
        header = list(bvd_results_df.columns)
        for item in range(len(header)):
            if '201' in header[item]:  # to determine the first column with financil
                first_fin_data_row = item
                break
        return first_fin_data_row