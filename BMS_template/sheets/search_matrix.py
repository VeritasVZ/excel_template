# -*- coding: utf-8 -*-
import re
from openpyxl.worksheet.worksheet import *
from BMS_template.template_properties import Properties,Styles
from BMS_template.bms_template import BvDResultsSheet

class SearchMatrix:

    def search_matrix_header(lang):

        first_header_row_types = {
            'BvD Database data':('Дані за бази даних','Данные из базы данных','BvD Database data'),
            'Additional criteria':('Додаткові критерії','Дополнительные критерии','Additional criteria')}
        subheader_type = {
            'criterion 1': ('Відбір за виручкою','Отбор по выручке','Revenue threshold'),
            'criterion 2': ('Відбір за прибутком','Отбор по прибыли','Net Profit threshold'),
            'criterion 3': ('Вибірка для перевірки в інтернеті','Выборка для проверки в интернете','Web search'),
            'criterion 4': ('Відбір за доступністю інформації','Отбор по доступности информации','Information availability'),
            'criterion 5': ('Відбір за предметом операцій','Отбор по предмету операций','Product comparability'),
            'criterion 6': ('Відбір за зіставністю функцій','Отбор по функциональному профилю','Functions comparability'),
            'criterion 7': ('Відбір за наявністю додаткової незіставної діяльності','Отбор по наличию дополнительной несопоставимой деятельности','Additional uncomparable functions'),
            'criterion 8': ('Відбір за критерієм незалежності','Отбор по критерию независимости','Independence check'),
            'criterion 9': ('Вибірка зіставних компаній','Выборка сопоставимых компаний','Final comparable set')
        }
        if lang == 'UA':
            header_type = 0
        elif lang == 'RU':
            header_type = 1
        elif lang == 'EU':
            header_type = 2
        else:
            print('Wrong language')
        first_header_row = [first_header_row_types['BvD Database data'][header_type], first_header_row_types['Additional criteria'][header_type]]
        subheader_row = []
        for criterion_no in range(0,len(list(subheader_type.keys()))+1):
            criterion = subheader_type.get(list(subheader_type.keys())[criterion_no-1])
            subheader_row.append(criterion[header_type])
        return first_header_row,subheader_row

    def copy_from_results(bms_template, bvd_results_df):
        sheet_names = bms_template.get_sheet_names()
        bvd_results_sheet = bms_template.get_sheet_by_name(sheet_names[0])
        matrix_sheet = bms_template.get_sheet_by_name(sheet_names[4])

        matrix_sheet.row_dimensions[1].height = 25
        matrix_sheet['A1'] = SearchMatrix.search_matrix_header(Properties.lang_from_template_type(bms_template))[0][0]
        Styles.header_style_dark_blue(matrix_sheet['A1'])
        matrix_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        matrix_sheet['F1'] = SearchMatrix.search_matrix_header(Properties.lang_from_template_type(bms_template))[0][1]
        Styles.header_style_green(matrix_sheet['F1'])
        matrix_sheet.merge_cells(start_row=1, start_column=6, end_row=1, end_column=15)

        for col in range(1,6):
            cell_value = bvd_results_sheet.cell(column=col, row=1).value
            cell_in_template = matrix_sheet.cell(column=col, row=2, value=cell_value)
            Styles.header_style_dark_blue(cell_in_template)
            for row in range(3,len(tuple(bvd_results_sheet.rows))+1):
                cell_value = bvd_results_sheet.cell(column=col, row=row-1).value
                cell_in_template = matrix_sheet.cell(column=col, row=row, value=cell_value)
                Styles.basic_style(cell_in_template)
        Properties.define_col_width(matrix_sheet,(1,6),3)

        for col in range (6,16):
            cell_value = SearchMatrix.search_matrix_header(Properties.lang_from_template_type(bms_template))[1][col-6]
            cell_in_template = matrix_sheet.cell(column=col, row=2, value=cell_value)
            Properties.define_col_width(matrix_sheet,(6,16),2)
            Styles.header_style_green(cell_in_template)

        for col in range(16,len(SearchMatrix.parsing_fin_data_header(bvd_results_df)[0])+15,3):
            cell_value = str(SearchMatrix.parsing_fin_data_header(bvd_results_df)[1][int((col-16)/3)])
            cell_in_template = matrix_sheet.cell(column=col,row=1,value = cell_value)
            Styles.header_style_dark_blue(cell_in_template)
            matrix_sheet.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+2)
            for subheader_col in range(col,col+3):
                subheader_cell_value = str(SearchMatrix.parsing_fin_data_header(bvd_results_df)[2][int(subheader_col-col)])
                subheader_cell = matrix_sheet.cell(column=subheader_col,row=2,value = subheader_cell_value)
                Styles.header_style_dark_blue(subheader_cell)
                for row in range(3, len(tuple(bvd_results_sheet.rows)) + 1):
                    cell_value = bvd_results_sheet.cell(column=subheader_col-1, row=row - 1).value
                    cell_in_template = matrix_sheet.cell(column=subheader_col, row=row, value=cell_value)
                    Styles.basic_style(cell_in_template)
                    Properties.number_to_format(matrix_sheet,column=subheader_col,row=row,format_str='0.0')
            Properties.define_col_width(matrix_sheet, (16, len(SearchMatrix.parsing_fin_data_header(bvd_results_df)[0])+15), 3)

        print('results copied to search matrix')
        return bms_template


    def parsing_fin_data_header(bvd_results_df):

        header = list(bvd_results_df.columns)
        fin_data_header = header[BvDResultsSheet.first_fin_data_row(bvd_results_df):]
        years=[]
        statements=[]
        for item in fin_data_header:
            statement = str(re.findall(r'.+',item))
            year = str(re.findall('20\d+',item)).replace('[','').replace(']','').replace("'",'')
            if year not in years:
                years.append(year)
            statement = str(statement).replace('(Rate at last closing date)','').replace('[','').replace(']','').replace("'",'')
            statement = ''.join(statement)
            statement = statement[:-6]
            if statement not in statements:
                statements.append(statement)
        return fin_data_header,statements,years
