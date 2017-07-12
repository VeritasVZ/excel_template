import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import *
from openpyxl.worksheet.worksheet import *
from BMS_template.template_properties import Properties
from BMS_template.sheets.strategy import StrategySheet
from BMS_template.sheets.bvd_results import BvDResultsSheet
from BMS_template.sheets.manual_steps import ManualStep
from BMS_template.sheets.search_matrix import SearchMatrix


class Template:

    def create_template(lang): #template lang: 'UA', 'RU', 'EU'
        bms_template = Workbook()
        sheet_names = Properties.template_type(lang)
        sheet_bvd_results = bms_template.active
        sheet_bvd_results.title = str(sheet_names['sheet_1'])
        for sheet_no in range(2,len(sheet_names)+1):
            new_sheet = bms_template.create_sheet(title=str(sheet_names['sheet_'+str(sheet_no)]))
        print('template created')
        return bms_template

    def build_template(bms_template,bvd_results_df,strategy,lang):
        bms_template = StrategySheet.write_search_strategy(bms_template=bms_template, search_strategy=strategy, lang=lang)
        bms_template = BvDResultsSheet.write_bvd_results(bms_template=bms_template, bvd_results_df=bvd_results_df)
        bms_template = ManualStep.steps_table(bms_template=bms_template)
        bms_template = SearchMatrix.copy_from_results(bms_template=bms_template,bvd_results_df=bvd_results_df)
        return bms_template


class InputParsing:

    def get_strategy_from_export(file_name):
        ruslana_export = load_workbook(file_name)
        search_strategy = ruslana_export['Стратегия поиска']
        results = ruslana_export['Результаты']
        print('strategy received from input file')
        return search_strategy

    def get_results_from_export(file_name):  # takes excel sheet from openpyxl Workbook with list of companies and their data from BvD DB
        bvd_results_df = pd.read_excel(file_name, 'Результаты', index_col=None, na_values=['NA'])
        print('bvd_results received from input file')
        return bvd_results_df

